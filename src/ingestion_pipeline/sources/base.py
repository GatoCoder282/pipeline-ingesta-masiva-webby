"""Input adapters. They only parse; they do not apply Webby business rules."""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Protocol

from ingestion_pipeline.domain.models import ParsedTable


class SourceAdapter(Protocol):
    def read(self, path: Path, *, sheet_name: str | None = None) -> ParsedTable: ...


def _unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers, start=1):
        clean = header.strip() or f"column_{index}"
        seen[clean] = seen.get(clean, 0) + 1
        result.append(clean if seen[clean] == 1 else f"{clean}__{seen[clean]}")
    return result


class CsvSource:
    def read(self, path: Path, *, sheet_name: str | None = None) -> ParsedTable:
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:  # pragma: no cover - the fallback always handles single-byte input
            text = raw.decode("utf-8", errors="replace")
        sample = text[:4096]
        first_line = next((line for line in sample.splitlines() if line.strip()), "")
        delimiters = ",;\t|"
        delimiter = max(delimiters, key=first_line.count)
        if first_line.count(delimiter) == 0:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=delimiters).delimiter
            except csv.Error:
                delimiter = ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        if not rows:
            return ParsedTable([], [], path)
        headers = _unique_headers(rows[0])
        width = len(headers)
        body = [row[:width] + [None] * max(0, width - len(row)) for row in rows[1:]]
        return ParsedTable(headers, body, path)


class XlsxSource:
    def read(self, path: Path, *, sheet_name: str | None = None) -> ParsedTable:
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"No se pudo leer el Excel {path}: {exc}") from exc
        try:
            if sheet_name and sheet_name not in workbook.sheetnames:
                raise ValueError(f"La hoja `{sheet_name}` no existe. Hojas: {workbook.sheetnames}")
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        if not rows:
            return ParsedTable([], [], path, sheet_name)
        headers = _unique_headers(["" if value is None else str(value) for value in rows[0]])
        width = len(headers)
        body = [row[:width] + [None] * max(0, width - len(row)) for row in rows[1:]]
        return ParsedTable(headers, body, path, worksheet.title)


def read_table(path: Path, *, sheet_name: str | None = None) -> ParsedTable:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Archivo de entrada no encontrado: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return XlsxSource().read(path, sheet_name=sheet_name)
    if suffix in {".csv", ".txt", ".tsv"}:
        return CsvSource().read(path, sheet_name=sheet_name)
    raise ValueError("Formato no soportado. Usa .csv, .tsv, .txt, .xlsx o .xlsm.")
